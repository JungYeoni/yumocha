"""TF-IDF 영역분류 결과를 수작업 검토용 Excel로 내보낸다."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from scripts.consolidate_2021_area_labels import MAJOR_BY_SUBCATEGORY
from src.modeling.similarity_grouping import (
    CONTENT_GROUP_COLUMN,
    CONTENT_GROUP_SCORE_COLUMN,
    NAME_GROUP_COLUMN,
    NAME_GROUP_SCORE_COLUMN,
)

REVIEW_STATUSES = ["미검토", "확정", "수정", "보류"]
REVIEW_COLUMNS = [
    "연도",
    "지역",
    "원본행",
    "세부사업명",
    "주요내용_정제",
    "예측_대영역",
    "예측_세부영역",
    "예측_신뢰도",
    "저신뢰_검토대상",
    "검토_대영역",
    "검토_세부영역",
    "검토상태",
    "검토메모",
    NAME_GROUP_COLUMN,
    NAME_GROUP_SCORE_COLUMN,
    CONTENT_GROUP_COLUMN,
    CONTENT_GROUP_SCORE_COLUMN,
]
REVIEW_KEY_COLUMNS = ["연도", "지역", "원본행"]
TRANSFER_COLUMNS = ["검토_세부영역", "검토상태", "검토메모"]
# 이미 사람이 검토를 끝낸 행(예: 2021-2024 확정 라벨)을 신규 예측과 함께
# 한 워크북에 담을 때, 있으면 그대로 쓰고 없으면 예측값·기본값으로
# 채우는 선택적 컬럼들이다. 프레임에 없으면 조용히 건너뛴다.
PREFILL_COLUMNS = ["검토_대영역", "검토_세부영역", "검토상태", "검토메모"]
NOTE_COLUMN = "명칭_내용_불일치_복합대응"
BUDGET_COLUMNS = ["당해예산", "전년도예산"]
SOURCE_LABEL_COLUMN = "자료구분"
OPTIONAL_TRAILING_COLUMNS = [NOTE_COLUMN, *BUDGET_COLUMNS, SOURCE_LABEL_COLUMN]


def validate_review_source(frame: pd.DataFrame) -> None:
    """검토용 Excel 생성에 필요한 예측 열을 확인한다."""
    required = [
        "연도",
        "지역",
        "원본행",
        "세부사업명",
        "주요내용_정제",
        "예측_대영역",
        "예측_세부영역",
        "예측_신뢰도",
        "저신뢰_검토대상",
    ]
    missing = [column for column in required if column not in frame.columns]
    if missing:
        raise ValueError(f"검토용 Excel 필수 열 누락: {missing}")
    if frame.empty:
        raise ValueError("검토할 예측 데이터가 비어 있습니다.")
    if frame[["연도", "지역", "원본행", "예측_세부영역"]].isna().any().any():
        raise ValueError("검토 키 또는 예측 라벨에 결측이 있습니다.")


def _normalized_review_key(values: tuple[object, object, object]) -> tuple[int, str, int]:
    """Excel에서 숫자로 읽힌 검토 키를 비교 가능한 형식으로 정규화한다."""
    year, region, original_row = values
    if year is None or region is None or original_row is None:
        raise ValueError("검토 키에 결측이 있습니다.")
    return int(float(year)), str(region).strip(), int(float(original_row))


def _header_positions(sheet) -> dict[str, int]:
    """검토 시트의 헤더명별 1-based 열 위치를 반환한다."""
    positions = {
        str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value is not None
    }
    required = [*REVIEW_KEY_COLUMNS, "예측_세부영역", *TRANSFER_COLUMNS]
    missing = [column for column in required if column not in positions]
    if missing:
        raise ValueError(f"검토 Excel 필수 열 누락: {missing}")
    return positions


def transfer_review_progress(
    existing_review_path: Path,
    refreshed_review_path: Path,
    output_path: Path,
    *,
    excluded_years: tuple[int, ...] = (),
) -> dict[str, int]:
    """기존 검토값을 키가 같은 새 예측 Excel에 이관한다."""
    from openpyxl import load_workbook

    existing_workbook = load_workbook(existing_review_path, data_only=False)
    refreshed_workbook = load_workbook(refreshed_review_path, data_only=False)
    existing_sheet = existing_workbook["영역분류검토"]
    refreshed_sheet = refreshed_workbook["영역분류검토"]
    existing_positions = _header_positions(existing_sheet)
    refreshed_positions = _header_positions(refreshed_sheet)

    progress_by_key: dict[tuple[int, str, int], tuple[object, object, object]] = {}
    excluded_rows = 0
    for row_number in range(2, existing_sheet.max_row + 1):
        raw_key = tuple(
            existing_sheet.cell(row_number, existing_positions[column]).value
            for column in REVIEW_KEY_COLUMNS
        )
        if all(value is None for value in raw_key):
            continue
        key = _normalized_review_key(raw_key)
        predicted = existing_sheet.cell(row_number, existing_positions["예측_세부영역"]).value
        reviewed = existing_sheet.cell(row_number, existing_positions["검토_세부영역"]).value
        status = existing_sheet.cell(row_number, existing_positions["검토상태"]).value
        memo = existing_sheet.cell(row_number, existing_positions["검토메모"]).value
        has_progress = (
            status not in (None, "", "미검토")
            or reviewed not in (None, "", predicted)
            or memo not in (None, "")
        )
        if not has_progress:
            continue
        if key[0] in excluded_years:
            excluded_rows += 1
            continue
        if key in progress_by_key:
            raise ValueError(f"기존 검토 Excel에 중복 키가 있습니다: {key}")
        progress_by_key[key] = (reviewed, status, memo)

    refreshed_rows: dict[tuple[int, str, int], int] = {}
    for row_number in range(2, refreshed_sheet.max_row + 1):
        raw_key = tuple(
            refreshed_sheet.cell(row_number, refreshed_positions[column]).value
            for column in REVIEW_KEY_COLUMNS
        )
        if all(value is None for value in raw_key):
            continue
        key = _normalized_review_key(raw_key)
        if key in refreshed_rows:
            raise ValueError(f"새 검토 Excel에 중복 키가 있습니다: {key}")
        refreshed_rows[key] = row_number

    missing_keys = sorted(set(progress_by_key).difference(refreshed_rows))
    if missing_keys:
        raise ValueError(f"새 검토 Excel에서 기존 검토 키를 찾을 수 없습니다: {missing_keys[:5]}")

    for key, values in progress_by_key.items():
        row_number = refreshed_rows[key]
        for column, value in zip(TRANSFER_COLUMNS, values, strict=True):
            if column == "검토상태" and value in (None, ""):
                continue
            refreshed_sheet.cell(row_number, refreshed_positions[column]).value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    refreshed_workbook.save(output_path)
    existing_workbook.close()
    refreshed_workbook.close()
    return {
        "이관건수": len(progress_by_key),
        "제외건수": excluded_rows,
        "새파일행수": len(refreshed_rows),
    }


def _prefill(row: object, column: str, default: object) -> object:
    """프레임에 해당 열이 있고 값이 결측이 아니면 그 값을, 아니면 기본값을 쓴다."""
    value = getattr(row, column, None)
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    return value


def create_review_workbook(
    frame: pd.DataFrame, output_path: Path, *, preserve_order: bool = False
) -> Path:
    """예측값을 초기 검토값으로 채운 드롭다운 Excel을 생성한다.

    ``preserve_order``가 True면 전달된 프레임의 행 순서를 그대로 쓴다.
    `similarity_grouping.assign_similarity_groups`처럼 이미 원하는 순서로
    정렬해 둔 프레임을 넘길 때 사용한다. False(기본값)면 기존과 동일하게
    연도·지역·원본행 순으로 정렬한다.

    이미 검토가 끝난 행(예: 2021-2024 확정 라벨)을 신규 예측과 함께 한
    워크북에 담을 때는, ``검토_대영역``·``검토_세부영역``·``검토상태``·
    ``검토메모`` 열을 프레임에 미리 채워서 넘기면 그 값을 그대로 쓰고
    (결측인 행만 예측값·"미검토"·빈 값으로 기본 채움), ``PREFILL_COLUMNS``
    자체가 없으면 기존과 동일하게 전부 새로 채운다. 예산(``당해예산``·
    ``전년도예산``), 명칭불일치·복합대응 메모(``명칭_내용_불일치_복합대응``),
    자료 출처 구분(``자료구분``)도 프레임에 있으면 뒤쪽 열에 추가된다.
    """
    validate_review_source(frame)
    ordered = (
        frame.reset_index(drop=True)
        if preserve_order
        else frame.sort_values(["연도", "지역", "원본행"]).reset_index(drop=True)
    )

    present_optional = [column for column in OPTIONAL_TRAILING_COLUMNS if column in ordered.columns]
    columns = [*REVIEW_COLUMNS, *present_optional]

    workbook = Workbook()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    review_sheet = workbook.active
    review_sheet.title = "영역분류검토"
    label_sheet = workbook.create_sheet("라벨목록")

    label_sheet.append(["세부영역", "대영역"])
    for subcategory, major in MAJOR_BY_SUBCATEGORY.items():
        label_sheet.append([subcategory, major])
    label_sheet.append([])
    label_sheet.append(["검토상태"])
    for status in REVIEW_STATUSES:
        label_sheet.append([status])

    review_sheet.append(columns)
    for row in ordered.itertuples(index=False):
        review_대영역 = _prefill(row, "검토_대영역", row.예측_대영역)
        review_세부영역 = _prefill(row, "검토_세부영역", row.예측_세부영역)
        review_상태 = _prefill(row, "검토상태", "미검토")
        review_메모 = _prefill(row, "검토메모", None)
        optional_values = []
        for column in present_optional:
            value = _prefill(row, column, None)
            if column in BUDGET_COLUMNS and value is not None:
                value = float(value)
            optional_values.append(value)

        review_sheet.append(
            [
                row.연도,
                row.지역,
                row.원본행,
                row.세부사업명,
                None if pd.isna(row.주요내용_정제) else row.주요내용_정제,
                row.예측_대영역,
                row.예측_세부영역,
                float(row.예측_신뢰도),
                bool(row.저신뢰_검토대상),
                review_대영역,
                review_세부영역,
                review_상태,
                review_메모,
                getattr(row, NAME_GROUP_COLUMN, None),
                getattr(row, NAME_GROUP_SCORE_COLUMN, None),
                getattr(row, CONTENT_GROUP_COLUMN, None),
                getattr(row, CONTENT_GROUP_SCORE_COLUMN, None),
                *optional_values,
            ]
        )

    last_row = review_sheet.max_row
    last_column_letter = get_column_letter(len(columns))
    subcategory_count = len(MAJOR_BY_SUBCATEGORY)
    status_start = subcategory_count + 4
    status_end = status_start + len(REVIEW_STATUSES) - 1

    subcategory_validation = DataValidation(
        type="list",
        formula1=f"'라벨목록'!$A$2:$A${subcategory_count + 1}",
        allow_blank=False,
    )
    status_validation = DataValidation(
        type="list",
        formula1=f"'라벨목록'!$A${status_start}:$A${status_end}",
        allow_blank=False,
    )
    review_sheet.add_data_validation(subcategory_validation)
    review_sheet.add_data_validation(status_validation)
    subcategory_validation.add(f"K2:K{last_row}")
    status_validation.add(f"L2:L{last_row}")

    for row_number in range(2, last_row + 1):
        review_sheet.cell(row=row_number, column=10).value = (
            f"=IFERROR(VLOOKUP(K{row_number},'라벨목록'!$A$2:$B$"
            f'{subcategory_count + 1},2,FALSE),"")'
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    low_confidence_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in review_sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in ("J", "K", "L", "M"):
        for cell in review_sheet[column][1:]:
            cell.fill = editable_fill

    review_sheet.conditional_formatting.add(
        f"A2:{last_column_letter}{last_row}",
        FormulaRule(formula=["$I2=TRUE"], fill=low_confidence_fill),
    )
    review_sheet.conditional_formatting.add(
        f"A2:{last_column_letter}{last_row}",
        FormulaRule(
            formula=['$L2="수정"'],
            fill=PatternFill("solid", fgColor="E2F0D9"),
        ),
    )

    review_sheet.freeze_panes = "D2"
    review_sheet.auto_filter.ref = f"A1:{last_column_letter}{last_row}"
    review_sheet.sheet_view.showGridLines = False
    review_sheet.row_dimensions[1].height = 28

    widths = {
        "A": 10,
        "B": 9,
        "C": 11,
        "D": 32,
        "E": 55,
        "F": 22,
        "G": 27,
        "H": 13,
        "I": 17,
        "J": 22,
        "K": 27,
        "L": 12,
        "M": 35,
        "N": 18,
        "O": 20,
        "P": 22,
        "Q": 23,
    }
    optional_widths = {
        NOTE_COLUMN: 40,
        "당해예산": 14,
        "전년도예산": 14,
        SOURCE_LABEL_COLUMN: 16,
    }
    for offset, column in enumerate(present_optional):
        widths[get_column_letter(18 + offset)] = optional_widths.get(column, 18)
    for column, width in widths.items():
        review_sheet.column_dimensions[column].width = width
    budget_columns = {
        get_column_letter(18 + offset)
        for offset, column in enumerate(present_optional)
        if column in BUDGET_COLUMNS
    }
    for row in review_sheet.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        row[7].number_format = "0.000"
        for index in (14, 16):
            row[index].number_format = "0.000"
        for cell in row:
            if cell.column_letter in budget_columns:
                cell.number_format = "#,##0"
    wrap_columns = {"D", "E", "M"}
    if NOTE_COLUMN in present_optional:
        wrap_columns.add(get_column_letter(18 + present_optional.index(NOTE_COLUMN)))
    for column in wrap_columns:
        for cell in review_sheet[column][1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in label_sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    label_sheet.column_dimensions["A"].width = 30
    label_sheet.column_dimensions["B"].width = 24
    label_sheet.sheet_state = "hidden"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
