"""재정팀 최신 검토본에 유사사업·시도·연도순 검토 시트를 추가한다."""

from __future__ import annotations

import argparse
from copy import copy, deepcopy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

KEY_COLUMNS = ["연도", "지역", "원본행"]
GROUP_COLUMNS = [
    "유사사업그룹ID",
    "유사그룹행수",
    "최근접유사도",
    "최근접사업명",
    "당해예산(백만원)",
    "전년도예산(백만원)",
]
MAX_AUTO_HOLD_ROWS = 1_000
REGION_ORDER = [
    "서울",
    "부산",
    "대구",
    "인천",
    "광주",
    "대전",
    "울산",
    "세종",
    "경기",
    "강원",
    "충북",
    "충남",
    "전북",
    "전남",
    "경북",
    "경남",
    "제주",
]
SHEET_NAMES = {
    "유사사업별 전체검토": "작업용_유사사업순",
    "전체검토": "백업_원본순서",
    "모아보기": "참고_국제해외교류",
    "모아보기..2": "참고_행사축제",
    "라벨목록": "설정_라벨목록",
    "유사사업명검토": "참고_이전유사사업",
    "통합QA": "QA_최초통합점검",
}


def _key(year: object, region: object, source_row: object) -> tuple[int, str, str]:
    return int(float(year)), str(region).strip(), str(int(float(source_row)))


def _read_lookup(path: Path) -> pd.DataFrame:
    frame = pd.read_csv(path, encoding="utf-8-sig", low_memory=False)
    required = {
        *KEY_COLUMNS,
        "유사사업그룹ID",
        "유사그룹행수",
        "최근접유사도",
        "최근접사업명",
        "당해예산",
        "전년도예산",
    }
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError(f"유사사업 파일 필수 컬럼 누락: {sorted(missing)}")
    if frame.duplicated(KEY_COLUMNS).any():
        raise ValueError("유사사업 파일의 연도·지역·원본행 키가 중복됩니다.")
    return frame


def _rename_sheets(workbook) -> None:
    """시트 역할이 드러나도록 이름과 수식 참조를 함께 갱신한다."""
    old_label_ref = "'라벨목록'!"
    new_label_ref = "'설정_라벨목록'!"
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = cell.value.replace(old_label_ref, new_label_ref)
    for defined_name in workbook.defined_names.values():
        if defined_name.attr_text:
            defined_name.attr_text = defined_name.attr_text.replace(old_label_ref, new_label_ref)
    for old_name, new_name in SHEET_NAMES.items():
        if old_name in workbook.sheetnames:
            workbook[old_name].title = new_name


def _replace_hold_sheet(workbook, source_sheet) -> None:
    """정적 보류 사본을 작업 시트에 연동되는 자동 목록으로 교체한다."""
    if "보류" in workbook.sheetnames:
        del workbook["보류"]
    if "자동_보류목록" in workbook.sheetnames:
        del workbook["자동_보류목록"]
    hold_sheet = workbook.create_sheet("자동_보류목록", 1)
    for column in range(1, source_sheet.max_column + 1):
        source_cell = source_sheet.cell(1, column)
        target_cell = hold_sheet.cell(1, column, source_cell.value)
        target_cell._style = copy(source_cell._style)
        target_cell.alignment = copy(source_cell.alignment)
        letter = get_column_letter(column)
        hold_sheet.column_dimensions[letter].width = source_sheet.column_dimensions[letter].width
        hold_sheet.column_dimensions[letter].hidden = source_sheet.column_dimensions[letter].hidden
    helper_column = source_sheet.max_column + 1
    helper_letter = get_column_letter(helper_column)
    hold_sheet.cell(1, helper_column, "보류행_자동추출번호")
    hold_sheet.column_dimensions[helper_letter].hidden = True
    source_status_range = "'작업용_유사사업순'!$L$2:$L$57980"
    for row in range(2, MAX_AUTO_HOLD_ROWS + 2):
        hold_sheet.cell(row, helper_column).value = (
            f"=IFERROR(AGGREGATE(15,6,(ROW({source_status_range})-"
            f"ROW('작업용_유사사업순'!$L$2)+1)/({source_status_range}=\"보류\"),"
            f'ROWS(${helper_letter}$2:{helper_letter}{row})),"")'
        )
        for column in range(1, source_sheet.max_column + 1):
            letter = get_column_letter(column)
            cell = hold_sheet.cell(row, column)
            cell.value = (
                f'=IF(${helper_letter}{row}="","",INDEX('
                f"'작업용_유사사업순'!{letter}$2:{letter}$57980,${helper_letter}{row}))"
            )
            cell._style = copy(source_sheet.cell(2, column)._style)
            cell.alignment = copy(source_sheet.cell(2, column).alignment)
    hold_sheet.freeze_panes = "A2"
    hold_sheet.auto_filter.ref = f"A1:AA{MAX_AUTO_HOLD_ROWS + 1}"
    hold_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def build_grouped_workbook(source_path: Path, lookup_path: Path, output_path: Path) -> None:
    lookup = _read_lookup(lookup_path)
    lookup_by_key = {
        _key(row.연도, row.지역, row.원본행): row for row in lookup.itertuples(index=False)
    }

    workbook = load_workbook(source_path)
    source = workbook["전체검토"]
    if source.max_row != 57_980:
        raise ValueError(f"전체검토는 헤더 포함 57,980행이어야 합니다: {source.max_row}")

    target_name = "유사사업별 전체검토"
    if target_name in workbook.sheetnames:
        del workbook[target_name]
    target = workbook.copy_worksheet(source)
    target.title = target_name
    target.freeze_panes = source.freeze_panes
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.data_validations = deepcopy(source.data_validations)
    target.conditional_formatting = ConditionalFormattingList()
    for rules in source.conditional_formatting._cf_rules.values():
        for rule in rules:
            target.conditional_formatting.add("A2:AA57980", deepcopy(rule))

    header_by_name = {
        source.cell(1, column).value: column for column in range(1, source.max_column + 1)
    }
    key_indexes = [header_by_name[name] for name in KEY_COLUMNS]
    source_data_columns = 19
    records = []
    hyperlink_by_key: dict[tuple[int, str, str], dict[int, object]] = {}
    for row_number in range(2, source.max_row + 1):
        values = [
            source.cell(row_number, column).value for column in range(1, source_data_columns + 1)
        ]
        row_key = _key(*(values[index - 1] for index in key_indexes))
        if row_key not in lookup_by_key:
            raise ValueError(f"유사사업·예산 연결 실패: {row_key}")
        links = {
            column: copy(source.cell(row_number, column).hyperlink)
            for column in range(1, source_data_columns + 1)
            if source.cell(row_number, column).hyperlink
        }
        if links:
            hyperlink_by_key[row_key] = links
        records.append((row_key, values, lookup_by_key[row_key]))

    if len({record[0] for record in records}) != 57_979:
        raise ValueError("전체검토의 연도·지역·원본행 키가 중복됩니다.")

    region_rank = {region: index for index, region in enumerate(REGION_ORDER)}
    records.sort(
        key=lambda item: (
            0 if int(item[2].유사그룹행수) > 1 else 1,
            str(item[2].유사사업그룹ID),
            region_rank.get(item[0][1], len(REGION_ORDER)),
            item[0][0],
            int(item[0][2]),
        )
    )

    for cell in target[1]:
        if cell.column <= source_data_columns:
            continue
        if cell.column > 21:
            cell.value = None
    start_column = 22
    header_style = copy(source.cell(1, 1)._style)
    text_style = copy(source.cell(2, 4)._style)
    number_style = copy(source.cell(2, 3)._style)
    for offset, header in enumerate(GROUP_COLUMNS):
        cell = target.cell(1, start_column + offset, header)
        cell._style = copy(header_style)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    target.cell(1, start_column + 4).comment = Comment(
        "원본 문서의 계획예산 값입니다. 예산 행 연결 전수 감사(#73) 완료 전까지 미검증 값으로 사용하세요.",
        "이정연",
    )
    target.cell(1, start_column + 5).comment = copy(target.cell(1, start_column + 4).comment)

    for target_row, (row_key, values, info) in enumerate(records, start=2):
        for column, value in enumerate(values, start=1):
            cell = target.cell(target_row, column)
            cell.value = value
            cell.hyperlink = None
        target.cell(
            target_row,
            header_by_name["검토_대영역"],
            f"=IFERROR(VLOOKUP(K{target_row},'라벨목록'!$A$2:$B$19,2,FALSE),\"\")",
        )
        for column, hyperlink in hyperlink_by_key.get(row_key, {}).items():
            target.cell(target_row, column).hyperlink = copy(hyperlink)

        group_values = [
            str(info.유사사업그룹ID),
            int(info.유사그룹행수),
            float(info.최근접유사도),
            str(info.최근접사업명),
            None if pd.isna(info.당해예산) else float(info.당해예산),
            None if pd.isna(info.전년도예산) else float(info.전년도예산),
        ]
        for offset, value in enumerate(group_values):
            cell = target.cell(target_row, start_column + offset)
            cell.value = value
            cell._style = copy(text_style if offset in (0, 3) else number_style)
            if offset == 2:
                cell.number_format = "0.000"
            elif offset in (4, 5):
                cell.number_format = "#,##0.###"
            if offset == 3:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [17, 13, 14, 38, 18, 18]
    for offset, width in enumerate(widths):
        target.column_dimensions[get_column_letter(start_column + offset)].width = width
    # 원본의 우측 요약 수식은 유지하되 그룹 검토 화면에서는 숨겨
    # 유사사업·예산 컬럼이 바로 이어 보이도록 한다.
    target.column_dimensions["T"].hidden = True
    target.column_dimensions["U"].hidden = True
    target.auto_filter.ref = (
        f"A1:{get_column_letter(start_column + len(GROUP_COLUMNS) - 1)}{target.max_row}"
    )

    workbook._sheets.remove(target)
    workbook._sheets.insert(0, target)
    _replace_hold_sheet(workbook, target)
    _rename_sheets(workbook)
    workbook.active = 0
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    base = Path("data/interim/영역분류_라벨링/재정팀_검토본/통합작업")
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("2021포함_전체57979건_영역분류_검토.xlsx"),
    )
    parser.add_argument(
        "--lookup",
        type=Path,
        default=base / "전체_57979건_유사사업명_검토용.csv",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=base / "2021포함_전체57979건_유사사업별_시도연도순_예산포함.xlsx",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_grouped_workbook(args.source, args.lookup, args.output)
    print(f"유사사업별 검토 XLSX 생성: {args.output}")


if __name__ == "__main__":
    main()
