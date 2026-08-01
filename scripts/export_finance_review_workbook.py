"""통합 CSV를 재정팀이 이어서 검토할 수 있는 XLSX로 내보낸다."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_HEADERS = [
    "연도",
    "지역",
    "원본행",
    "세부사업명",
    "주요내용_정제",
    "예측_대영역",
    "예측_세부영역",
    "예측_신뢰도",
    "저신뢰_검토대상",
    "검토_대영역",
    "검토_세부영역",
    "검토상태",
    "명칭-내용 불일치 / 원본PDF 동일 여부",
    "검토메모",
    "라벨출처",
    "기존검토본_주요내용_정제",
    "최신정제문_갱신여부",
    "#72_정제변경",
    "#72_재확인대상",
]
SIMILAR_HEADERS = [
    "유사사업그룹ID",
    "유사그룹행수",
    "최근접유사도",
    "최근접사업명",
    "연도",
    "지역",
    "원본행",
    "세부사업명",
    "주요내용_정제",
    "검토_대영역",
    "검토_세부영역",
    "검토상태",
    "#72_정제변경",
    "#72_재확인대상",
]
DISPLAY_HEADERS = {
    "#72_정제변경": "정제문 수정 여부",
}


def _read_csv(path: Path) -> pd.DataFrame:
    frame = pd.read_csv(
        path,
        encoding="utf-8-sig",
        keep_default_na=False,
        dtype={"원본행": "string"},
        low_memory=False,
    )
    for column in ("예측_신뢰도", "최근접유사도"):
        if column in frame:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame


def _excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def _write_main_sheet(sheet, master: pd.DataFrame, label_last_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="999999")
    data_border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_alignment = Alignment(vertical="top")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    for column, header in enumerate(MAIN_HEADERS, start=1):
        cell = sheet.cell(1, column, DISPLAY_HEADERS.get(header, header))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = data_border

    rows = []
    for record in master[MAIN_HEADERS].itertuples(index=False, name=None):
        rows.append([_excel_value(value) for value in record])
    for row_index, record in enumerate(rows, start=2):
        for column_index, value in enumerate(record, start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.alignment = top_alignment
            cell.border = data_border
        sheet.cell(
            row_index,
            10,
            f"=IFERROR(VLOOKUP(K{row_index},'라벨목록'!$A$2:$B${label_last_row},2,FALSE),\"\")",
        )

    last_row = len(master) + 1
    sheet.auto_filter.ref = f"A1:S{last_row}"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    widths = {
        "A": 9,
        "B": 9,
        "C": 11,
        "D": 38,
        "E": 65,
        "F": 20,
        "G": 23,
        "H": 13,
        "I": 14,
        "J": 20,
        "K": 23,
        "L": 12,
        "M": 28,
        "N": 34,
        "O": 24,
        "P": 55,
        "Q": 18,
        "R": 15,
        "S": 17,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column in ("D", "E", "M", "N", "O", "P"):
        for cell in sheet[column][1:]:
            cell.alignment = wrap_alignment

    sheet.data_validations.dataValidation = []
    label_validation = DataValidation(
        type="list",
        formula1="ReviewLabels",
        allow_blank=False,
    )
    status_validation = DataValidation(
        type="list",
        formula1='"미검토,확정,수정,보류"',
        allow_blank=False,
    )
    sheet.add_data_validation(label_validation)
    sheet.add_data_validation(status_validation)
    label_validation.add(f"K2:K{last_row}")
    status_validation.add(f"L2:L{last_row}")

    sheet.conditional_formatting._cf_rules.clear()
    sheet.conditional_formatting.add(
        f"A2:S{last_row}",
        FormulaRule(
            formula=["$R2=TRUE"],
            fill=PatternFill("solid", fgColor="FFF4CCCC"),
            stopIfTrue=True,
        ),
    )
    sheet.conditional_formatting.add(
        f"A2:S{last_row}",
        FormulaRule(
            formula=['$L2="수정"'],
            fill=PatternFill("solid", fgColor="FFFFF2CC"),
        ),
    )
    sheet.conditional_formatting.add(
        f"A2:S{last_row}",
        FormulaRule(
            formula=["$I2=TRUE"],
            fill=PatternFill("solid", fgColor="FFDDEBF7"),
        ),
    )
    sheet.title = "전체검토"


def _style_table_sheet(sheet, headers: list[str], row_count: int) -> None:
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E1F2")
    for cell in sheet[1]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_count + 1}"
    sheet.sheet_view.showGridLines = False


def _write_similarity_sheet(workbook, similarity: pd.DataFrame) -> None:
    if "유사사업명검토" in workbook.sheetnames:
        del workbook["유사사업명검토"]
    sheet = workbook.create_sheet("유사사업명검토")
    sheet.append([DISPLAY_HEADERS.get(header, header) for header in SIMILAR_HEADERS])
    for record in similarity[SIMILAR_HEADERS].itertuples(index=False, name=None):
        sheet.append([_excel_value(value) for value in record])
    _style_table_sheet(sheet, SIMILAR_HEADERS, len(similarity))
    widths = [16, 13, 14, 38, 9, 9, 11, 38, 65, 20, 23, 12, 15, 17]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_qa_sheet(workbook, qa: pd.DataFrame, master: pd.DataFrame) -> None:
    if "통합QA" in workbook.sheetnames:
        del workbook["통합QA"]
    sheet = workbook.create_sheet("통합QA")
    sheet.append(["검사항목", "결과", "기대값", "판정"])
    for item, result, expected in qa[["검사항목", "결과", "기대값"]].itertuples(
        index=False, name=None
    ):
        sheet.append([item, result, expected, "통과" if result == expected else "확인 필요"])
    sheet.append([])
    sheet.append(["검토상태", "행수"])
    for status, count in master["검토상태"].value_counts().items():
        sheet.append([status, int(count)])
    sheet.append([])
    sheet.append(["추가 확인", "행수"])
    sheet.append(["기존 검토본 대비 최신 정제문 갱신", int(master["최신정제문_갱신여부"].sum())])
    sheet.append(["정제문 수정 행", int(master["#72_정제변경"].sum())])
    sheet.append(["#72 기존 검토 완료 행 재확인", int(master["#72_재확인대상"].sum())])
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16
    sheet.freeze_panes = "A2"
    for row in (1, len(qa) + 3, len(qa) + 9):
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)


def export_workbook(
    source_workbook: Path,
    master_path: Path,
    similarity_path: Path,
    qa_path: Path,
    output_path: Path,
) -> None:
    master = _read_csv(master_path)
    similarity = _read_csv(similarity_path)
    qa = _read_csv(qa_path)
    source = load_workbook(source_workbook, read_only=True, data_only=True)
    label_rows = list(source["라벨목록"].iter_rows(min_row=1, max_col=2, values_only=True))
    source.close()
    while label_rows and not any(value not in (None, "") for value in label_rows[-1]):
        label_rows.pop()
    if not label_rows or tuple(label_rows[0]) != ("세부영역", "대영역"):
        raise ValueError("라벨목록의 첫 두 컬럼은 `세부영역`, `대영역` 순서여야 합니다.")
    if len(label_rows) < 2:
        raise ValueError("라벨목록에 사용할 라벨이 없습니다.")

    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = "전체검토"
    label_sheet = workbook.create_sheet("라벨목록")
    for row in label_rows:
        label_sheet.append(list(row))

    label_last_row = len(label_rows)
    _write_main_sheet(main_sheet, master, label_last_row)
    _write_similarity_sheet(workbook, similarity)
    _write_qa_sheet(workbook, qa, master)
    workbook.defined_names.add(
        DefinedName("ReviewLabels", attr_text=f"'라벨목록'!$A$2:$A${label_last_row}")
    )
    workbook["라벨목록"].sheet_state = "hidden"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    base = Path("data/interim/영역분류_라벨링/재정팀_검토본")
    work = base / "통합작업"
    parser.add_argument(
        "--source-workbook",
        type=Path,
        default=base / "현재검토본/2016_2020_2022_2024_TFIDF_영역분류_검토_v2_2016복구반영.xlsx",
    )
    parser.add_argument(
        "--master",
        type=Path,
        default=work / "전체_57979건_원본순서.csv",
    )
    parser.add_argument(
        "--similarity",
        type=Path,
        default=work / "전체_57979건_유사사업명_검토용.csv",
    )
    parser.add_argument("--qa", type=Path, default=work / "통합_QA.csv")
    parser.add_argument(
        "--output",
        type=Path,
        default=work / "2021포함_전체57979건_영역분류_검토.xlsx",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    export_workbook(
        args.source_workbook,
        args.master,
        args.similarity,
        args.qa,
        args.output,
    )
    print(f"검토용 XLSX 생성: {args.output}")


if __name__ == "__main__":
    main()
