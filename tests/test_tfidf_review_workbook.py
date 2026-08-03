"""TF-IDF 영역분류 수작업 검토 Excel 테스트."""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.modeling.tfidf_review_workbook import (
    create_review_workbook,
    transfer_review_progress,
)
from src.modeling.similarity_grouping import (
    CONTENT_GROUP_COLUMN,
    CONTENT_GROUP_SCORE_COLUMN,
    NAME_GROUP_COLUMN,
    NAME_GROUP_SCORE_COLUMN,
    assign_similarity_groups,
)


def _prediction_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "연도": [2020, 2019],
            "지역": ["서울", "부산"],
            "원본행": [1, 2],
            "세부사업명": ["청년 취업", "아이 돌봄"],
            "주요내용_정제": ["고용 지원", pd.NA],
            "예측_대영역": ["1. 경제·고용·주거", "2. 가족·생활"],
            "예측_세부영역": ["1-1. 고용여건", "2-1. 돌봄 여건"],
            "예측_신뢰도": [0.2, 0.9],
            "저신뢰_검토대상": [True, False],
        }
    )


def test_create_review_workbook_adds_dropdowns_and_formulas(tmp_path):
    path = tmp_path / "review.xlsx"

    create_review_workbook(_prediction_frame(), path)

    workbook = load_workbook(path, data_only=False)
    sheet = workbook["영역분류검토"]
    assert sheet["D2"].value == "아이 돌봄"
    assert sheet["E2"].value is None
    assert sheet["K2"].value == "2-1. 돌봄 여건"
    assert sheet["A3"].value == 2020
    assert sheet["B3"].value == "서울"
    assert sheet["C3"].value == 1
    assert sheet["L2"].value == "미검토"
    assert sheet["J2"].value.startswith("=IFERROR(VLOOKUP(")
    validations = {
        str(validation.sqref): validation.formula1
        for validation in sheet.data_validations.dataValidation
    }
    assert validations["K2:K3"] == "'라벨목록'!$A$2:$A$13"
    assert validations["L2:L3"] == "'라벨목록'!$A$16:$A$19"
    assert workbook["라벨목록"].sheet_state == "hidden"


def test_create_review_workbook_preserve_order_keeps_input_row_order(tmp_path):
    path = tmp_path / "review.xlsx"
    frame = _prediction_frame()  # 기본 정렬이면 2019/부산이 먼저 옴

    create_review_workbook(frame, path, preserve_order=True)

    workbook = load_workbook(path, data_only=False)
    sheet = workbook["영역분류검토"]
    assert sheet["A2"].value == 2020
    assert sheet["B2"].value == "서울"
    assert sheet["A3"].value == 2019
    assert sheet["B3"].value == "부산"


def test_create_review_workbook_exposes_similarity_group_columns(tmp_path):
    path = tmp_path / "grouped-review.xlsx"
    grouped = assign_similarity_groups(_prediction_frame())

    create_review_workbook(grouped, path, preserve_order=True)

    workbook = load_workbook(path, data_only=False)
    sheet = workbook["영역분류검토"]
    assert [sheet.cell(1, column).value for column in range(14, 18)] == [
        NAME_GROUP_COLUMN,
        NAME_GROUP_SCORE_COLUMN,
        CONTENT_GROUP_COLUMN,
        CONTENT_GROUP_SCORE_COLUMN,
    ]
    assert sheet["N2"].value
    assert sheet.auto_filter.ref == "A1:Q3"


def test_create_review_workbook_rejects_missing_prediction_columns(tmp_path):
    with pytest.raises(ValueError, match="필수 열 누락"):
        create_review_workbook(pd.DataFrame({"연도": [2020]}), tmp_path / "review.xlsx")


def test_transfer_review_progress_preserves_new_predictions_and_excludes_year(tmp_path):
    existing_path = tmp_path / "existing.xlsx"
    refreshed_path = tmp_path / "refreshed.xlsx"
    output_path = tmp_path / "transferred.xlsx"
    existing = _prediction_frame()
    refreshed = existing.copy()
    refreshed["예측_세부영역"] = ["1-2. 주거여건", "2-2. 여가 인프라"]
    refreshed["예측_대영역"] = ["1. 경제·고용·주거", "2. 가족·생활"]
    create_review_workbook(existing, existing_path)
    create_review_workbook(refreshed, refreshed_path)

    workbook = load_workbook(existing_path)
    sheet = workbook["영역분류검토"]
    sheet["K2"] = "지표체계 외"
    sheet["L2"] = "수정"
    sheet["M2"] = "기존 검토 메모"
    sheet["K3"] = "4-2. 사회적 가치관"
    sheet["L3"] = "확정"
    workbook.save(existing_path)
    workbook.close()

    summary = transfer_review_progress(
        existing_path,
        refreshed_path,
        output_path,
        excluded_years=(2020,),
    )

    assert summary == {"이관건수": 1, "제외건수": 1, "새파일행수": 2}
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["영역분류검토"]
    assert sheet["G2"].value == "2-2. 여가 인프라"
    assert sheet["K2"].value == "지표체계 외"
    assert sheet["L2"].value == "수정"
    assert sheet["M2"].value == "기존 검토 메모"
    assert sheet["G3"].value == "1-2. 주거여건"
    assert sheet["K3"].value == "1-2. 주거여건"
    assert sheet["L3"].value == "미검토"
    assert sheet["J2"].value.startswith("=IFERROR(VLOOKUP(")
    workbook.close()
